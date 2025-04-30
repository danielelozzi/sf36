import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

def create_full_sf36_excel(filename="SF36_completo.xlsx"):
    wb = Workbook()
    
    # === Fase 1: Foglio Input ===
    ws_input = wb.active
    ws_input.title = "Input"

    domande = [
        "Age", "Sex",
        "Q1", "Q2", "Q3a", "Q3b", "Q3c", "Q3d", "Q3e", "Q3f", "Q3g", "Q3h", "Q3i", "Q3j",
        "Q4a", "Q4b", "Q4c", "Q4d",
        "Q5a", "Q5b", "Q5c",
        "Q6", "Q7", "Q8",
        "Q9a", "Q9b", "Q9c", "Q9d", "Q9e", "Q9f", "Q9g", "Q9h", "Q9i",
        "Q10", "Q11a", "Q11b", "Q11c", "Q11d"
    ]
    ws_input.append(domande)

    # === Fase 2: Foglio Output ===
    ws_out = wb.create_sheet("Output")
    headers = ["Scala", "Punteggio 0–100", "Z-score (USA)", "T-score (ITA)", "ISF/PCS weight", "ISM/MCS weight"]
    ws_out.append(headers)

    scale_map = {
        'PF': ["Q3a", "Q3b", "Q3c", "Q3d", "Q3e", "Q3f", "Q3g", "Q3h", "Q3i", "Q3j"],
        'RP': ["Q4a", "Q4b", "Q4c", "Q4d"],
        'RE': ["Q5a", "Q5b", "Q5c"],
        'VT': ["Q9a", "Q9e", "Q9g", "Q9i"],
        'MH': ["Q9b", "Q9c", "Q9d", "Q9f", "Q9h"],
        'SF': ["Q6", "Q10"],
        'BP': ["Q7", "Q8"],
        'GH': ["Q1", "Q11a", "Q11b", "Q11c", "Q11d"]
    }

    us_mean = {'PF':84.52, 'RP':81.19, 'BP':75.49, 'GH':72.21, 'VT':61.05, 'SF':83.60, 'RE':81.29, 'MH':74.84}
    us_sd = {'PF':22.89, 'RP':33.80, 'BP':23.56, 'GH':20.17, 'VT':20.87, 'SF':22.38, 'RE':33.03, 'MH':18.01}

    pcs_weight = {'PF':0.42, 'RP':0.35, 'BP':0.32, 'GH':0.25, 'VT':0.03, 'SF':-0.01, 'RE':-0.19, 'MH':-0.22}
    mcs_weight = {'PF':-0.23, 'RP':-0.12, 'BP':-0.10, 'GH':-0.02, 'VT':0.24, 'SF':0.27, 'RE':0.43, 'MH':0.49}

    # Fila per ogni scala
    for i, (scale, items) in enumerate(scale_map.items(), start=2):
        ws_out.cell(row=i, column=1, value=scale)
        letters = [get_column_letter(domande.index(q)+1) for q in items]
        formula_sum = "+".join([f"Input!{l}2" for l in letters])
        min_sum = len(items)
        max_sum = len(items)*5  # assumiamo scala 1–5 per default
        score_formula = f"=(({formula_sum}) - {min_sum}) / ({max_sum - min_sum}) * 100"
        ws_out.cell(row=i, column=2, value=score_formula)

        # Z-score
        z_formula = f"=(B{i}-{us_mean[scale]})/{us_sd[scale]}"
        ws_out.cell(row=i, column=3, value=z_formula)

        # Placeholder per T-score ITA (può essere modificato manualmente o da script esterno)
        ws_out.cell(row=i, column=4, value=f"=B{i}")  # copia valore come esempio

        # Pesi ISF/ISM
        ws_out.cell(row=i, column=5, value=pcs_weight.get(scale, 0))
        ws_out.cell(row=i, column=6, value=mcs_weight.get(scale, 0))

    # === Fase 3: Indici Sintetici ===
    ws_out.append(["ISF/PCS", f"=SUMPRODUCT(B2:B9,E2:E9)"])
    ws_out.append(["ISM/MCS", f"=SUMPRODUCT(B2:B9,F2:F9)"])

    # === Fase 4: Grafici dinamici ===
    def add_chart(title, col, pos):
        chart = BarChart()
        chart.title = title
        chart.y_axis.title = "Punteggio"
        chart.x_axis.title = "Scala"
        data = Reference(ws_out, min_col=col, min_row=1, max_row=9)
        cats = Reference(ws_out, min_col=1, min_row=2, max_row=9)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_out.add_chart(chart, pos)

    add_chart("Punteggi 0–100", 2, "H2")
    add_chart("Z-score (USA)", 3, "H18")
    add_chart("T-score (ITA)", 4, "H34")

    # === Salvataggio ===
    wb.save(filename)
    print(f"File Excel creato: {filename}")

create_full_sf36_excel()
